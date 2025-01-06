import os
import time
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import shutil
import matplotlib.colors as mcolors
import matplotlib.cm as cm
from tqdm import tqdm
import matplotlib.animation as animation

plt.switch_backend('Agg')

# Constants
ACTIVATION_FN = torch.relu
LOSS_FN = nn.MSELoss()

# Helper functions
def save_model_predictions(model, function, device, epoch, x_vals, all_predictions):
    """Saves the model's predictions for a given epoch in a dictionary.

    Args:
        model (torch.nn.Module): The model to be evaluated.
        function (str): The function the model is approximating ('f1', 'f2', etc.).
        device (torch.device): The device for computation (e.g., 'cpu' or 'cuda').
        epoch (int): The current epoch.
        x_vals (np.ndarray): Input data as a NumPy array.
        all_predictions (dict): Dictionary to store model predictions.

    Returns:
        None
    """
    model.eval()
    if function == "f3":
        x1, x2 = x_vals[:, 0], x_vals[:, 1]
        x1 = torch.tensor(x1, dtype=torch.float32).view(-1, 1).to(device)
        x2 = torch.tensor(x2, dtype=torch.float32).view(-1, 1).to(device)
        with torch.no_grad():
            y_pred = model(torch.cat((x1, x2), dim=1)).cpu().numpy()
    else:
        x = torch.tensor(x_vals, dtype=torch.float32).view(-1, 1).to(device)
        with torch.no_grad():
            y_pred = model(x).cpu().numpy()
    all_predictions[epoch] = y_pred

def create_animation_data(function):
    """Prepares data for animation.

    Args:
        function (str): The function to animate ('f1', 'f2', etc.).

    Returns:
        x_vals (np.ndarray): Input data for the animation.
        y_true (np.ndarray): True function values for the animation.
    """
    if function == "f3":
        size = 50
        x1 = np.linspace(-0.5, 0.5, size).astype(np.float32)
        x2 = np.linspace(-0.5, 0.5, size).astype(np.float32)
        x1_flat, x2_flat = np.meshgrid(x1, x2)
        x_vals = np.column_stack((x1_flat.ravel(), x2_flat.ravel()))
        y_true = np.array([1 if x1_i + np.abs(x2_i) ** (3 / 2) >= 0 else 0 for x1_i, x2_i in x_vals])
    else:
        if function in ["f1", "f5", "f6"]:
            a, b = 0, 1
        elif function == "f4":
            a, b = -1, 1
        else:
            a, b = -0.5, 0.5
        x_vals = np.linspace(a, b, 400).astype(np.float32)  # Use 400 points for animation
        y_true = get_function_values(x_vals, function)

    return x_vals, y_true

def create_animation(exp_dir, function, x_vals, y_true, frames, all_predictions, interval=33.33):
    """Creates and saves an animation of model fitting over epochs at 60 FPS.

    Args:
        exp_dir (str): Directory to save the animation.
        function (str): The function to animate ('f1', 'f2', etc.).
        x_vals (np.ndarray): Input data for the animation.
        y_true (np.ndarray): True function values for the animation.
        frames (list): List of epochs for which predictions were saved.
        all_predictions (dict): Dictionary with predictions for each epoch.
        interval (float): Time interval between frames in milliseconds.

    Returns:
        None
    """
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.set_title(f"Training Progress for {function}")

    # Position the epoch text in the top-right corner of the plot
    epoch_text = ax.text(0.95, 0.95, '', transform=ax.transAxes,
                         fontsize=12, verticalalignment='top', horizontalalignment='right', color='red')

    if function == "f3":
        # Setup for f3: Color scheme based on predictions
        x1, x2 = x_vals[:, 0], x_vals[:, 1]
        x2_line = np.linspace(-0.5, 0.5, 1000)
        x1_line = -np.abs(x2_line) ** (3 / 2)
        ax.plot(x2_line, x1_line, color='blue', label='$x_1 = |x_2|^{3/2}$')
        sc = ax.scatter(x2, x1, c=y_true, cmap=cm.RdYlGn_r, norm=mcolors.Normalize(vmin=0, vmax=1), s=50)
        
        # Add colorbar
        cbar = plt.colorbar(sc, ax=ax, orientation='vertical')
        cbar.set_label('Difference (|Model Prediction - True Value|)')

        def update(frame):
            current_epoch = frames[frame]  # Get the actual epoch number from frames
            y_pred = all_predictions[current_epoch]
            differences = np.abs(y_pred.flatten() - y_true)
            sc.set_array(differences)
            epoch_text.set_text(f'Epoch: {current_epoch}')
            return sc, epoch_text

    else:
        # Setup for other functions
        line1, = ax.plot(x_vals, y_true, label='True Function')
        line2, = ax.plot([], [], label='Model Prediction')

        def update(frame):
            current_epoch = frames[frame]  # Get the actual epoch number from frames
            y_pred = all_predictions[current_epoch]
            line2.set_data(x_vals, y_pred.flatten())
            epoch_text.set_text(f'Epoch: {current_epoch}')
            return line2, epoch_text

    ani = animation.FuncAnimation(fig, update, frames=range(len(frames)), interval=interval, blit=True)
    ani.save(os.path.join(exp_dir, f'animation_{function}.mp4'), writer='ffmpeg')
    plt.close(fig)

def calculate_computation_units(hidden_layers):
    """Calculates the number of computation units for given hidden layers.

    Args:
        hidden_layers (list): List of neuron counts in hidden layers.

    Returns:
        int: Total number of computation units.
    """
    return sum(hidden_layers)

def calculate_weights(layers, allow_non_adjacent=False):
    """Calculates the total number of weights in a network.

    Args:
        layers (list): List of neuron counts in all layers.
        allow_non_adjacent (bool): If True, includes non-adjacent connections.

    Returns:
        int: Total number of weights.
    """
    if allow_non_adjacent:
        weights = sum(layers[i] * sum(layers[i + 1:]) for i in range(len(layers) - 1))
    else:
        weights = sum(layers[i] * layers[i + 1] for i in range(len(layers) - 1))
    return weights + sum(layers[1:])

def save_to_excel(filename, data, sheet_name="Sheet1"):
    """Saves given data to an Excel file.

    Args:
        filename (str): Name of the Excel file.
        data (list): Data to save (list of rows).
        sheet_name (str): Name of the worksheet.

    Returns:
        None
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    for row in data:
        ws.append(row)
    
    wb.save(filename)

def plot_function_3(x_vals, y_true, y_pred, exp_dir, title):
    """Creates a scatter plot for function f3 and saves it as an image.

    Args:
        x_vals (np.ndarray): Input data (x1 and x2) for function f3.
        y_true (np.ndarray): True function values.
        y_pred (np.ndarray): Predicted values by the model.
        exp_dir (str): Directory to save the plot.
        title (str): Title of the plot.

    Returns:
        None
    """
    x1, x2 = x_vals[:, 0], x_vals[:, 1]
    differences = np.abs(y_pred - y_true)
    norm = mcolors.Normalize(vmin=0, vmax=1)
    cmap = cm.RdYlGn_r

    plt.figure(figsize=(10, 10))
    sc = plt.scatter(x2, x1, c=differences, cmap=cmap, norm=norm, s=50)

    x2_line = np.linspace(-0.5, 0.5, 1000)
    x1_line = -np.abs(x2_line) ** (3 / 2)
    plt.plot(x2_line, x1_line, color='blue', label='$x_1 = |x_2|^{3/2}$')

    plt.colorbar(sc, label='Difference (|Prediction - True Value|)')
    plt.xlabel('$x_2$')
    plt.ylabel('$x_1$')
    plt.title(title)
    plt.legend()
    plt.grid(True)
    plt.savefig(os.path.join(exp_dir, 'function_plot_f3.png'))
    plt.close()

def plot_and_save(x_vals, y_true, y_pred, exp_dir, title, function):
    """Creates and saves a plot based on the function and given values.

    Args:
        x_vals (np.ndarray): Input data.
        y_true (np.ndarray): True function values.
        y_pred (np.ndarray): Predicted values by the model.
        exp_dir (str): Directory to save the plot.
        title (str): Title of the plot.
        function (str): Name of the function ('f1', 'f2', etc.).

    Returns:
        None
    """
    if function == "f3":
        plot_function_3(x_vals, y_true, y_pred, exp_dir, title)
    else:
        plt.figure()
        plt.plot(x_vals, y_true, label='True Function')
        plt.plot(x_vals, y_pred, label='Model Prediction')
        plt.xlabel('x')
        plt.ylabel('y')
        plt.title(title)
        plt.legend()
        plt.savefig(os.path.join(exp_dir, 'function_plot.png'))
        plt.close()

def save_loss_plot(train_losses, test_losses, exp_dir):
    """Saves a plot of training and test losses over epochs.

    Args:
        train_losses (list): List of tuples (epoch, training loss).
        test_losses (list): List of tuples (epoch, test loss, ...).
        exp_dir (str): Directory to save the plot.

    Returns:
        None
    """
    plt.figure()
    epochs_train, losses_train = zip(*[(epoch, loss) for epoch, loss in train_losses])
    epochs_test, losses_test = zip(*[(epoch, test_loss) for epoch, test_loss, _, _, _, _, _ in test_losses])
    best_losses_test = [min(losses_test[:i + 1]) for i in range(len(losses_test))]

    plt.plot(epochs_train, losses_train, label='Training Loss')
    plt.plot(epochs_test, losses_test, label='Test Loss')
    plt.plot(epochs_test, best_losses_test, linestyle='--', label='Best Test Loss')

    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.yscale('log')
    plt.title('Training and Test Loss over Epochs (Log Scale)')
    plt.legend()
    plt.savefig(os.path.join(exp_dir, 'loss_plot.png'))
    plt.close()

def save_experiment_data(exp_dir, data, headers):
    """Saves experiment data and loss trajectories in an Excel file.

    Args:
        exp_dir (str): Directory to save the Excel file.
        data (dict): Dictionary with experiment data.
        headers (list): List of headers for the Excel file.

    Returns:
        None
    """
    filename = os.path.join(exp_dir, f'experiment_data.xlsx')
    os.makedirs(exp_dir, exist_ok=True)
    
    values = [
        data['function'], data['optimizer_type'], data['lr'], ', '.join(map(str, data['layers'])), data['comp_units'],
        data['weights'], data['train_size'], data['test_size'], data['epochs'], data['loss_function'],
        data['best_test_loss'], data['total_train_time'], data['best_train_time'], data['non_zero_weights'],
        data['non_zero_weights_l1'], data['non_zero_weights_l2'], data['non_zero_weights_sup'],
        data['architecture'], str(data['early_stopping']), data['actual_epochs'], ', '.join(map(str, data['hidden_layers'])),
        data['best_l1_norm'], data['best_l2_norm'], data['best_sup_norm'],
        data['time_to_best_l1'], data['time_to_best_l2'], data['time_to_best_sup'],
        data['actual_parameters']
    ]
    
    test_loss_headers = ["Epoch", "Test Loss", "Training Time", "Best Test Loss", "L1 Norm", "L2 Norm", "Sup Norm"]
    train_loss_headers = ["Epoch", "Training Loss"]

    test_losses_transposed = list(zip(*[(epoch, test_loss, train_time, best_loss, l1, l2, sup) 
                                        for epoch, test_loss, train_time, best_loss, l1, l2, sup in data['test_losses']
                                        if (epoch + 1) % 50 == 0 or epoch == 0]))

    train_losses_transposed = list(zip(*[(epoch, loss) 
                                         for epoch, loss in data['train_losses']
                                         if (epoch + 1) % 50 == 0 or epoch == 0]))
    
    rows = [
        headers,
        values,
        [],
        test_loss_headers,
        *zip(*test_losses_transposed),
        [],
        train_loss_headers,
        *zip(*train_losses_transposed),
    ]
    
    save_to_excel(filename, rows, "Experiment Data")

# Model-related functions
class Net(nn.Module):
    """Defines a neural network with optional non-adjacent connections.

    Attributes:
        allow_non_adjacent (bool): Indicates if non-adjacent connections are allowed.
        layers (nn.ModuleList): List of standard layers (adjacent connections).
        depth (int): Number of layers in the network.
        non_adjacent_layers (nn.ModuleDict): Dictionary of non-adjacent layers.
    """
    def __init__(self, input_size, hidden_layers, output_size, allow_non_adjacent=False):
        """Initializes the network.

        Args:
            input_size (int): Size of the input vector.
            hidden_layers (list): List of neuron counts in hidden layers.
            output_size (int): Size of the output layer.
            allow_non_adjacent (bool): If True, non-adjacent connections are created.

        Returns:
            None
        """
        super(Net, self).__init__()
        self.allow_non_adjacent = allow_non_adjacent

        self.layers = self._create_standard_layers(input_size, hidden_layers, output_size)
        self.depth = len(self.layers)
        self.non_adjacent_layers = nn.ModuleDict()
        if allow_non_adjacent:
            self._create_non_adjacent_layers(input_size, hidden_layers, output_size)

    def _create_standard_layers(self, input_size, hidden_layers, output_size):
        """Creates standard layers enabling adjacent connections.

        Args:
            input_size (int): Size of the input vector.
            hidden_layers (list): List of neuron counts in hidden layers.
            output_size (int): Size of the output layer.

        Returns:
            nn.ModuleList: List of standard layers.
        """
        layers = nn.ModuleList()
        sizes = [input_size] + hidden_layers + [output_size]
        for in_size, out_size in zip(sizes[:-1], sizes[1:]):
            layers.append(nn.Linear(in_size, out_size))
        return layers

    def _create_non_adjacent_layers(self, input_size, hidden_layers, output_size):
        """Creates non-adjacent layers enabling non-consecutive connections.

        Args:
            input_size (int): Size of the input vector.
            hidden_layers (list): List of neuron counts in hidden layers.
            output_size (int): Size of the output layer.

        Returns:
            None
        """
        sizes = [input_size] + hidden_layers + [output_size]
        for i in range(len(sizes) - 2):
            for j in range(i + 2, len(sizes)):
                self.non_adjacent_layers[f'layer_{i}_to_{j}'] = nn.Linear(sizes[i], sizes[j])

    def forward(self, x):
        """Defines the forward pass of the network.

        Args:
            x (torch.Tensor): Input data.

        Returns:
            torch.Tensor: Output of the network.
        """
        outputs = [x]
        for i in range(self.depth - 1):
            y = torch.zeros_like(self.layers[i](outputs[i]))

            y += self.layers[i](outputs[i])

            if self.allow_non_adjacent:
                for j in range(i):
                    layer_key = f'layer_{j}_to_{i + 1}'
                    if layer_key in self.non_adjacent_layers:
                        y += self.non_adjacent_layers[layer_key](outputs[j])
                        
            y = ACTIVATION_FN(y)
            outputs.append(y)

        y = torch.zeros_like(self.layers[-1](outputs[-1]))
        y += self.layers[-1](outputs[-1])

        if self.allow_non_adjacent:
            for j in range(self.depth - 1):
                layer_key = f'layer_{j}_to_{self.depth}'
                if layer_key in self.non_adjacent_layers:
                    y += self.non_adjacent_layers[layer_key](outputs[j])

        return y

def train_one_epoch(model, device, x_train, y_train, optimizer):
    """Trains the model for one epoch.

    Args:
        model (torch.nn.Module): The model to be trained.
        device (torch.device): The device for computation.
        x_train (torch.Tensor): Training input data.
        y_train (torch.Tensor): Training output data.
        optimizer (torch.optim.Optimizer): Optimization algorithm.

    Returns:
        float: Value of the loss function after the epoch.
    """
    model.train()
    x_train, y_train = x_train.to(device), y_train.to(device)
    optimizer.zero_grad()
    loss = LOSS_FN(model(x_train), y_train)
    loss.backward()
    optimizer.step()
    return loss.item()

def evaluate_model(model, device, x_test, y_test):
    """Evaluates the model on the test data.

    Args:
        model (torch.nn.Module): The model to be evaluated.
        device (torch.device): The device for computation.
        x_test (torch.Tensor): Test input data.
        y_test (torch.Tensor): Test output data.

    Returns:
        float: Value of the loss function on the test data.
    """
    model.eval()
    x_test, y_test = x_test.to(device), y_test.to(device)
    with torch.no_grad():
        loss = LOSS_FN(model(x_test), y_test).item()
    return loss

def count_non_zero_weights(model):
    """Counts the number of non-zero weights in the model.

    Args:
        model (torch.nn.Module): The model to be analyzed.

    Returns:
        int: Number of non-zero weights.
    """
    return sum((param != 0).sum().item() for param in model.parameters())

def save_model(model, path):
    """Saves the state of the model to a file.

    Args:
        model (torch.nn.Module): The model to be saved.
        path (str): Path to the save file.

    Returns:
        None
    """
    torch.save(model.state_dict(), path)

# Data generation and processing
def generate_data(train_size, test_size, function):
    """Generates training and test data for the given function.

    Args:
        train_size (int): Number of training data points.
        test_size (int): Number of test data points.
        function (str): Name of the function ('f1', 'f2', etc.).

    Returns:
        x_train_tensor (torch.Tensor): Training input data.
        y_train_tensor (torch.Tensor): Training output data.
        x_test_tensor (torch.Tensor): Test input data.
        y_test_tensor (torch.Tensor): Test output data.
    """
    functions = {
        "f1": lambda x: np.cos(np.pi * x) / (np.pi ** 2),
        "f2": lambda x: np.abs(x) ** (3 / 2),
        "f3": lambda x1, x2: (x1 + np.abs(x2) ** (3 / 2)) >= 0,
        "f4": lambda x: np.abs(x) ** (7 / 2),
        "f5": lambda x: x ** 2,
        "f6": lambda x: np.cos(2 * np.pi * x),
    }
    if function in ["f1", "f5", "f6"]:
        boundaries = [0, 1]
    elif function == "f4":
        boundaries = [-1, 1]
    else:
        boundaries = [-0.5, 0.5]
    if function == "f3":
        return generate_multivariate_data(train_size, test_size, boundaries, functions[function])
    return generate_univariate_data(train_size, test_size, boundaries, functions[function])

def generate_univariate_data(train_size, test_size, boundaries, function):
    """Generates univariate data for training and testing.

    Args:
        train_size (int): Number of training data points.
        test_size (int): Number of test data points.
        boundaries (list): Range for random numbers [min, max].
        function (callable): Function to generate output data.

    Returns:
        x_train_tensor (torch.Tensor): Training input data.
        y_train_tensor (torch.Tensor): Training output data.
        x_test_tensor (torch.Tensor): Test input data.
        y_test_tensor (torch.Tensor): Test output data.
    """
    x_train = np.random.uniform(boundaries[0], boundaries[1], train_size).astype(np.float32)
    x_test = np.random.uniform(boundaries[0], boundaries[1], test_size).astype(np.float32)
    y_train, y_test = function(x_train), function(x_test)
    x_train_tensor = torch.tensor(x_train).view(-1, 1)
    x_test_tensor = torch.tensor(x_test).view(-1, 1)
    y_train_tensor = torch.tensor(y_train).view(-1, 1)
    y_test_tensor = torch.tensor(y_test).view(-1, 1)
    return x_train_tensor, y_train_tensor, x_test_tensor, y_test_tensor

def generate_multivariate_data(train_size, test_size, boundaries, function):
    """Generates multivariate data for training and testing.

    Args:
        train_size (int): Number of training data points.
        test_size (int): Number of test data points.
        boundaries (list): Range for random numbers [min, max].
        function (callable): Function to generate output data.

    Returns:
        x_train_tensor (torch.Tensor): Training input data.
        y_train_tensor (torch.Tensor): Training output data.
        x_test_tensor (torch.Tensor): Test input data.
        y_test_tensor (torch.Tensor): Test output data.
    """
    x1_train = np.random.uniform(boundaries[0], boundaries[1], train_size).astype(np.float32)
    x2_train = np.random.uniform(boundaries[0], boundaries[1], train_size).astype(np.float32)
    x1_test = np.random.uniform(boundaries[0], boundaries[1], test_size).astype(np.float32)
    x2_test = np.random.uniform(boundaries[0], boundaries[1], test_size).astype(np.float32)
    y_train = np.array([function(x1, x2) for x1, x2 in zip(x1_train, x2_train)]).astype(np.float32)
    y_test = np.array([function(x1, x2) for x1, x2 in zip(x1_test, x2_test)]).astype(np.float32)
    x_train = np.column_stack((x1_train, x2_train))
    x_test = np.column_stack((x1_test, x2_test))
    x_train_tensor = torch.tensor(x_train).view(-1, 2)
    x_test_tensor = torch.tensor(x_test).view(-1, 2)
    y_train_tensor = torch.tensor(y_train).view(-1, 1)
    y_test_tensor = torch.tensor(y_test).view(-1, 1)
    return x_train_tensor, y_train_tensor, x_test_tensor, y_test_tensor

def compute_norm(model, device, function, norm_type):
    """Computes the specified norm for the model.

    Args:
        model (torch.nn.Module): The model to be evaluated.
        device (torch.device): The device for computation.
        function (str): Name of the function ('f1', 'f2', etc.).
        norm_type (str): Type of norm to compute ('sup', 'l1', 'l2').

    Returns:
        float: Value of the computed norm.
    """
    norm_funcs = {
        "sup": lambda y, y_pred, dx: np.max(np.abs(y - y_pred)),
        "l1": lambda y, y_pred, dx: np.sum(np.abs(y - y_pred)) * dx,
        "l2": lambda y, y_pred, dx: np.sqrt(np.sum((y - y_pred) ** 2) * dx),
    }
    x, y, dx = prepare_data_for_norm_calculation(function)
    if function == "f3":
        x = torch.tensor(x).to(device)
    else:
        x = torch.tensor(x).view(-1, 1).to(device)
    with torch.no_grad():
        y_pred = model(x).cpu().detach().numpy().flatten()
    return norm_funcs[norm_type](y, y_pred, dx)

def prepare_data_for_norm_calculation(function, plot=False):
    """Prepares data for calculating norms.

    Args:
        function (str): Name of the function ('f1', 'f2', etc.).
        plot (bool): Indicates if the data is prepared for plotting.

    Returns:
        x (np.ndarray): Input data.
        y (np.ndarray): True function values.
        dx (float): Distance between data points.
    """
    if function == "f3":
        if plot:
            size = 50
        else:
            size = 100
        x1 = np.linspace(-0.5, 0.5, size).astype(np.float32)
        x2 = np.linspace(-0.5, 0.5, size).astype(np.float32)
        x1_flat, x2_flat = np.meshgrid(x1, x2)
        x = np.column_stack((x1_flat.ravel(), x2_flat.ravel()))
        y = np.array([1 if x1_i + np.abs(x2_i) ** (3 / 2) >= 0 else 0 for x1_i, x2_i in x])
        dx = (0.5 - (-0.5)) / len(x1) / len(x2)
    else:
        if function in ["f1", "f5", "f6"]:
            a, b = 0, 1
        elif function == "f4":
            a, b = -1, 1
        else:
            a, b = -0.5, 0.5
        x = np.linspace(a, b, 10000).astype(np.float32)
        y = get_function_values(x, function)
        dx = 1 / len(x)
    return x, y, dx

def get_function_values(x, function):
    """Gets function values for a given x and a function.

    Args:
        x (np.ndarray): Input data.
        function (str): Name of the function ('f1', 'f2', etc.).

    Returns:
        np.ndarray: Output data of the function.
    """
    functions = {
        "f1": lambda x: np.cos(np.pi * x) / (np.pi ** 2),
        "f2": lambda x: np.abs(x) ** (3 / 2),
        "f4": lambda x: np.abs(x) ** (7 / 2),
        "f5": lambda x: x ** 2,
        "f6": lambda x: np.cos(2 * np.pi * x),
    }
    return functions[function](x)

# Main logic for training and evaluation
def train_and_evaluate(model, device, x_train, y_train, x_test, y_test, optimizer, epochs, early_stopping, function, hidden_layers, exp_dir, is_first_run, animation=False):
    """Trains and evaluates the model, saves the best results, and tracks performance over epochs.

    Args:
        model (torch.nn.Module): The model to be trained.
        device (torch.device): The device for computation.
        x_train (torch.Tensor): Training input data.
        y_train (torch.Tensor): Training output data.
        x_test (torch.Tensor): Test input data.
        y_test (torch.Tensor): Test output data.
        optimizer (torch.optim.Optimizer): Optimization algorithm.
        epochs (int): Number of training epochs.
        early_stopping (tuple): Parameters for early stopping (window size, threshold).
        function (str): Name of the function ('f1', 'f2', etc.).
        hidden_layers (list): List of neuron counts in hidden layers.
        exp_dir (str): Directory to save experiment data.
        is_first_run (bool): Indicates if this is the first run (for outputs).
        animation (bool): If True, creates an animation.

    Returns:
        tuple: Contains training losses, test losses, best test losses, model states, times, and other metrics.
    """
    if is_first_run:
        print(f"Starting training for function {function} with hidden layers: {hidden_layers}")

    train_losses, test_losses = [], []
    best_test_loss = float('inf')
    best_l1_norm = float('inf')
    best_l2_norm = float('inf')
    best_sup_norm = float('inf')
    best_model_state = None
    best_model_state_l1 = None
    best_model_state_l2 = None
    best_model_state_sup = None
    time_to_best_loss, cumulative_train_time = 0, 0
    time_to_best_l1, time_to_best_l2, time_to_best_sup = 0, 0, 0
    es_window, es_threshold = early_stopping

    x_vals, y_true = create_animation_data(function)
    prediction_epochs = []
    all_predictions = {}

    progress_bar = tqdm(range(epochs), desc="Training", unit="epoch")
    start_time = time.time()
    for epoch in progress_bar:
        train_loss = train_one_epoch(model, device, x_train, y_train, optimizer)
        train_losses.append((epoch, train_loss))

        if (epoch + 1) % 10 == 0 or epoch == 0:
            cumulative_train_time = time.time() - start_time
            test_loss = evaluate_model(model, device, x_test, y_test)
            l1_norm = compute_norm(model, device, function, "l1")
            l2_norm = compute_norm(model, device, function, "l2")
            sup_norm = compute_norm(model, device, function, "sup")

            if l1_norm < best_l1_norm:
                best_l1_norm = l1_norm
                best_model_state_l1 = model.state_dict()
                time_to_best_l1 = cumulative_train_time
            if l2_norm < best_l2_norm:
                best_l2_norm = l2_norm
                best_model_state_l2 = model.state_dict()
                time_to_best_l2 = cumulative_train_time
            if sup_norm < best_sup_norm:
                best_sup_norm = sup_norm
                best_model_state_sup = model.state_dict()
                time_to_best_sup = cumulative_train_time
            if test_loss < best_test_loss:
                best_test_loss = test_loss
                best_model_state = model.state_dict()
                time_to_best_loss = cumulative_train_time

            test_losses.append((epoch, test_loss, cumulative_train_time, best_test_loss, l1_norm, l2_norm, sup_norm))

            if ((epoch + 1) % 100 == 0) or epoch == 0:
                save_model_predictions(model, function, device, epoch + 1, x_vals, all_predictions)
                prediction_epochs.append(epoch + 1)

            if (epoch + 1) % 50 == 0 or epoch == 0:
                progress_bar.set_postfix(train_loss=train_loss, test_loss=test_loss, best_test_loss=best_test_loss)

            if len(test_losses) > es_window / 10 and best_test_loss >= (1 - es_threshold / 100) * test_losses[int(-es_window / 10)][3]:
                print(f"Early stopping at epoch {epoch + 1}")
                break

    print(f"Training complete. Final epoch: {epoch + 1}/{epochs}, Best test loss: {best_test_loss}")

    if animation:
        create_animation(exp_dir, function, x_vals, y_true, prediction_epochs, all_predictions)

    return (train_losses, test_losses, best_test_loss, best_model_state, cumulative_train_time, 
            time_to_best_loss, epoch + 1, best_model_state_l1, best_model_state_l2, best_model_state_sup, 
            best_l1_norm, best_l2_norm, best_sup_norm, time_to_best_l1, time_to_best_l2, time_to_best_sup)

def run_experiment(hidden_layers, epochs, train_size, test_size, lr, optimizer_type, device, function="f5", num_experiments=5, network_type="MLP", architecture="unspecified", early_stopping=(10000, 10), animation=False):
    """Runs an experiment by training and evaluating a model for a given function and network architecture.

    Args:
        hidden_layers (list): List of neuron counts in the hidden layers.
        epochs (int): Number of training epochs.
        train_size (int): Number of training data points.
        test_size (int): Number of test data points.
        lr (float): Learning rate.
        optimizer_type (str): Type of optimization algorithm ('adam' or 'sgd').
        device (torch.device): Device for computation.
        function (str): Name of the function ('f1', 'f2', etc.).
        num_experiments (int): Number of experiments to run.
        network_type (str): Type of network ('MLP' or 'FF' for Feedforward).
        architecture (str): Label for the network architecture.
        early_stopping (tuple): Parameters for early stopping (window size, threshold).
        animation (bool): If True, an animation will be created.

    Returns:
        None
    """
    input_dimension = 2 if function == "f3" else 1
    layers = [input_dimension] + hidden_layers + [1]

    main_dir_name = create_experiment_directory_name(
        function, layers, hidden_layers, epochs, train_size, optimizer_type, lr, "MSELoss", network_type, architecture, early_stopping
    )

    if os.path.exists(main_dir_name):
        print(f"The directory {main_dir_name} already exists. An experiment with these parameters has already been conducted. Skipping.")
        return

    create_experiment_directories(main_dir_name, num_experiments)

    x_train, y_train, x_test, y_test = generate_data(train_size, test_size, function)

    best_norms = {"sup": float('inf'), "l1": float('inf'), "l2": float('inf'), "loss": float('inf')}
    best_models = {"sup": None, "l1": None, "l2": None, "loss": None}
    best_exp = {"sup": None, "l1": None, "l2": None, "loss": None}
    best_train_time = {"sup": 0, "l1": 0, "l2": 0, "loss": 0}
    best_non_zero_weights = {"sup": float('inf'), "l1": float('inf'), "l2": float('inf'), "loss": float('inf')}

    is_first_run = True
    for exp_num in range(1, num_experiments + 1):
        allow_non_adjacent = network_type == "FF"
        model = Net(x_train.shape[1], hidden_layers, y_train.shape[1], allow_non_adjacent).to(device)
        optimizer = optim.Adam(model.parameters(), lr=lr) if optimizer_type == 'adam' else optim.SGD(model.parameters(), lr=lr)

        exp_dir = os.path.join(main_dir_name, f"Exp{exp_num}")
        results = train_and_evaluate(
            model, device, x_train, y_train, x_test, y_test, optimizer, epochs, early_stopping, function, hidden_layers, exp_dir, is_first_run, animation
        )
        is_first_run = False

        (train_losses, test_losses, best_test_loss, best_model_state, total_train_time, 
         time_to_best_loss, actual_epochs, best_model_state_l1, best_model_state_l2, best_model_state_sup, 
         best_l1_norm, best_l2_norm, best_sup_norm, time_to_best_l1, time_to_best_l2, time_to_best_sup) = results

        actual_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad)

        if best_model_state_l1:
            save_model(model, os.path.join(exp_dir, f'best_model_l1_{best_l1_norm:.8f}.pth'))
        if best_model_state_l2:
            save_model(model, os.path.join(exp_dir, f'best_model_l2_{best_l2_norm:.8f}.pth'))
        if best_model_state_sup:
            save_model(model, os.path.join(exp_dir, f'best_model_sup_{best_sup_norm:.8f}.pth'))
        if best_model_state:
            save_model(model, os.path.join(exp_dir, f'best_model_loss_{best_test_loss:.8f}.pth'))

        non_zero_weights_l1 = sum(torch.count_nonzero(param).item() for param in best_model_state_l1.values())
        non_zero_weights_l2 = sum(torch.count_nonzero(param).item() for param in best_model_state_l2.values())
        non_zero_weights_sup = sum(torch.count_nonzero(param).item() for param in best_model_state_sup.values())
        non_zero_weights = sum(torch.count_nonzero(param).item() for param in best_model_state.values())

        if best_sup_norm < best_norms["sup"]:
            best_norms["sup"] = best_sup_norm
            best_models["sup"] = model.state_dict()
            best_exp["sup"] = exp_num
            best_train_time["sup"] = time_to_best_sup
            best_non_zero_weights["sup"] = non_zero_weights_sup
        if best_l1_norm < best_norms["l1"]:
            best_norms["l1"] = best_l1_norm
            best_models["l1"] = model.state_dict()
            best_exp["l1"] = exp_num
            best_train_time["l1"] = time_to_best_l1
            best_non_zero_weights["l1"] = non_zero_weights_l1
        if best_l2_norm < best_norms["l2"]:
            best_norms["l2"] = best_l2_norm
            best_models["l2"] = model.state_dict()
            best_exp["l2"] = exp_num
            best_train_time["l2"] = time_to_best_l2
            best_non_zero_weights["l2"] = non_zero_weights_l2
        if best_test_loss < best_norms["loss"]:
            best_norms["loss"] = best_test_loss
            best_models["loss"] = model.state_dict()
            best_exp["loss"] = exp_num
            best_train_time["loss"] = time_to_best_loss
            best_non_zero_weights["loss"] = non_zero_weights

        experiment_data = {
            'function': function,
            'optimizer_type': optimizer_type,
            'lr': lr,
            'layers': layers,
            'comp_units': calculate_computation_units(hidden_layers),
            'weights': calculate_weights(layers, allow_non_adjacent),
            'train_size': train_size,
            'test_size': test_size,
            'epochs': epochs,
            'loss_function': "MSELoss",
            'best_test_loss': best_test_loss,
            'total_train_time': total_train_time,
            'best_train_time': time_to_best_loss,
            'non_zero_weights': non_zero_weights,
            'non_zero_weights_l1': non_zero_weights_l1,
            'non_zero_weights_l2': non_zero_weights_l2,
            'non_zero_weights_sup': non_zero_weights_sup,
            'architecture': architecture,
            'early_stopping': early_stopping,
            'actual_epochs': actual_epochs,
            'hidden_layers': hidden_layers,
            'best_l1_norm': best_l1_norm,
            'best_l2_norm': best_l2_norm,
            'best_sup_norm': best_sup_norm,
            'time_to_best_l1': time_to_best_l1,
            'time_to_best_l2': time_to_best_l2,
            'time_to_best_sup': time_to_best_sup,
            'actual_parameters': actual_parameters,
            'train_losses': train_losses,
            'test_losses': test_losses,
        }

        headers = [
            "Function", "Optimizer", "Learning Rate", "Layers", "Computation Units", "Weights",
            "Train Size", "Test Size", "Epochs", "Loss Function", "Best Test Loss",
            "Total Training Time", "Training Time to Best Test Loss", "Non-zero Weights",
            "Non-zero Weights L1", "Non-zero Weights L2", "Non-zero Weights Sup",
            "Architecture", "Early Stopping", "Actual Epochs", "Hidden Layers",
            "Best L1 Norm", "Best L2 Norm", 
            "Best Supremum Norm", "Time to Best L1", "Time to Best L2", 
            "Time to Best Supremum", "Actual Parameters"
        ]

        save_experiment_data(exp_dir, experiment_data, headers)
        save_loss_plot(train_losses, test_losses, exp_dir)
        plot_norms(test_losses, exp_dir)

        plot = False
        if function == "f3":
            plot = True
        x_vals, y_true, _ = prepare_data_for_norm_calculation(function, plot)
        x = None
        if function == "f3":
            x = torch.tensor(x_vals).to(device)
        else:
            x = torch.tensor(x_vals).view(-1, 1).to(device)
        with torch.no_grad():
            y_pred = model(x).cpu().numpy().flatten()
        plot_and_save(x_vals, y_true, y_pred, exp_dir, f'{function} Function vs Model Prediction', function)

    for norm, model_state in best_models.items():
        if model_state:
            model = Net(x_train.shape[1], hidden_layers, y_train.shape[1], allow_non_adjacent)
            model.load_state_dict(model_state)
            save_model(model, os.path.join(main_dir_name, f'best_{norm}_Exp_{best_exp[norm]}_Value_{best_norms[norm]:.8f}.pth'))
            exp_filename = 'experiment_data.xlsx'
            src_file = os.path.join(main_dir_name, f"Exp{best_exp[norm]}", exp_filename)
            dst_file = os.path.join(main_dir_name, f'{norm.capitalize()}_Exp_{best_exp[norm]}.xlsx')
            shutil.copy(src_file, dst_file)

    update_summary_file(function, [function, layers, hidden_layers, epochs, train_size, test_size, optimizer_type, lr], actual_parameters, best_norms, architecture, best_train_time, early_stopping, best_non_zero_weights)

def plot_norms(test_losses, exp_dir):
    """Creates and saves plots of norms over epochs.

    Args:
        test_losses (list): List of test losses and norms over epochs.
        exp_dir (str): Directory to save the plots.

    Returns:
        None
    """
    epochs = [loss[0] for loss in test_losses]
    l1_norms = [loss[4] for loss in test_losses]
    l2_norms = [loss[5] for loss in test_losses]
    sup_norms = [loss[6] for loss in test_losses]

    for norm_values, norm_name in zip([l1_norms, l2_norms, sup_norms], ["L1 Norm", "L2 Norm", "Supremum Norm"]):
        plt.figure()
        best_norm_values = [min(norm_values[:i + 1]) for i in range(len(norm_values))]

        plt.plot(epochs, norm_values, label=f'{norm_name} Over Epochs')
        plt.plot(epochs, best_norm_values, linestyle='--', label=f'Best {norm_name}')

        plt.xlabel('Epochs')
        plt.ylabel(norm_name)
        plt.yscale('log')
        plt.title(f'{norm_name} over Epochs (Log Scale)')
        plt.legend()
        plt.savefig(os.path.join(exp_dir, f'{norm_name.lower().replace(" ", "_")}_plot.png'))
        plt.close()

def create_experiment_directories(main_dir_name, num_experiments):
    """Creates directories for the experiments.

    Args:
        main_dir_name (str): Main directory name.
        num_experiments (int): Number of experiments.

    Returns:
        None
    """
    os.makedirs(main_dir_name, exist_ok=True)
    for i in range(1, num_experiments + 1):
        os.makedirs(os.path.join(main_dir_name, f"Exp{i}"), exist_ok=True)

def create_experiment_directory_name(function, layers, hidden_layers, epochs, train_size, optimizer_type, lr, loss_function, network_type, architecture, early_stopping):
    """Creates a name for the experiment directory based on parameters.

    Args:
        function (str): Name of the function ('f1', 'f2', etc.).
        layers (list): List of neuron counts in all layers.
        hidden_layers (list): List of neuron counts in hidden layers.
        epochs (int): Number of training epochs.
        train_size (int): Number of training data points.
        optimizer_type (str): Type of optimization algorithm ('adam' or 'sgd').
        lr (float): Learning rate.
        loss_function (str): Loss function.
        network_type (str): Type of network ('MLP' or 'FF' for Feedforward).
        architecture (str): Label for the network architecture.
        early_stopping (tuple): Parameters for early stopping (window size, threshold).

    Returns:
        str: Name of the experiment directory.
    """
    comp_units = calculate_computation_units(hidden_layers)
    weights = calculate_weights(layers, network_type == "FF")
    early_stopping_str = f"ES_{early_stopping[0]}_{early_stopping[1]}"
    dir_name = (
        f"Arch_{architecture}_Layers_{len(layers)}_Units_{comp_units}_Weights_{weights}_"
        f"Opt_{optimizer_type}_LR_{lr}_Epochs_{epochs}_TrainSize_{train_size}_"
        f"Loss_{loss_function}_NetType_{network_type}_{early_stopping_str}"
    )
    return os.path.join(function, dir_name)

def update_summary_file(function, params, actual_parameters, best_norms, architecture, best_train_time, early_stopping, best_non_zero_weights):
    """Updates the summary file with the experiment results.

    Args:
        function (str): Name of the function ('f1', 'f2', etc.).
        params (list): List of experiment parameters.
        actual_parameters (int): Number of parameters trained in the model.
        best_norms (dict): Dictionary of best norms.
        architecture (str): Label for the network architecture.
        best_train_time (dict): Dictionary of training times to achieve best norms.
        early_stopping (tuple): Parameters for early stopping (window size, threshold).
        best_non_zero_weights (dict): Dictionary of non-zero weight counts.

    Returns:
        None
    """
    summary_file = f'experiment_summary_{function}.xlsx'
    
    if os.path.exists(summary_file):
        wb = load_workbook(summary_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = [
            "Function", "Layers", "Neurons per Layer", "Epochs", "Train Size", "Test Size",
            "Optimizer", "Learning Rate", "Best Supremum Norm", "Best L1 Norm", "Best L2 Norm", 
            "Best Test Loss", "Actual Parameters", "Architecture", "Early Stopping Window", "Early Stopping Threshold",
            "Train Time Best Sup", "Train Time Best L1", "Train Time Best L2", "Train Time Best Loss",
            "Non-zero Weights Sup", "Non-zero Weights L1", "Non-zero Weights L2", "Non-zero Weights Loss"
        ]
        ws.append(headers)

    function, layers, hidden_layers, epochs, train_size, test_size, optimizer_type, lr = params

    hidden_layers_str = ", ".join(map(str, hidden_layers))
    layers_str = ", ".join(map(str, layers))
    
    row = [
        function, layers_str, hidden_layers_str, epochs, train_size, test_size, optimizer_type, lr,
        best_norms["sup"], best_norms["l1"], best_norms["l2"], best_norms["loss"],
        actual_parameters, architecture, early_stopping[0], early_stopping[1],
        best_train_time["sup"], best_train_time["l1"], best_train_time["l2"], best_train_time["loss"],
        best_non_zero_weights["sup"], best_non_zero_weights["l1"], best_non_zero_weights["l2"], best_non_zero_weights["loss"]
    ]

    ws.append(row)
    wb.save(summary_file)

# Enable anomaly detection for PyTorch autograd
torch.autograd.set_detect_anomaly(True)
